from asyncio.windows_events import NULL
from numpy import double
import openpyxl
import csv
import os
import time
from enum import Enum
# import json

def SwitchType(s_Type):
    return {'integer': int, 'string': str, 'bool': bool, 'float': float, 'double': double}.get(s_Type, str)

def GetType(value):
    return type(value)


class DataType(Enum):
    INTEGER = 'integer'
    STRING = 'string'
    BOOL = 'bool'
    FLOAT = 'float'


class FileFormat:
    def __init__(self, fileName):
        # 파일 이름
        self.fileName = fileName
        # 삭제할 열의 내용
        self.del_ColName = ['//', 'comment']
        self.completeConvert = False
        self.isOpened = False
        self.convertList = []
        self.isTypename_Insheet_List = []
        self.delFirstrow = False
        self.savePath = ''

    def SetFileName(self, name):
        self.fileName = name
        if self.isOpened:
            self.wb.close()

        self.PrintLog("I'm on it open file...", 'black')
        try:
            self.wb = openpyxl.load_workbook(
                str(self.fileName), data_only=True)
            self.lastSaveTime = os.path.getmtime(name)
            self.PrintLog('Opening the file was successful', 'blue')
            self.isOpened = True
            # 숨어 있는 시트 제외
            self.sNames = self.DeleteHiddenSheet()
            self.convertList = self.sNames

        except:
            self.sNames = ''
            self.PrintLog('Error : Failed to open file', 'red')

        return self.sNames

    def SetSavePath(self, path, pathIsFilename=False):
        if pathIsFilename == True:
            if self.savePath != '':
                return self.savePath
            self.savePath = list(reversed(self.fileName.split("/")))
            self.savePath[0] = 'resultFolder'
            self.savePath = "/".join((reversed(self.savePath)))
        else:
            self.savePath = path

        return self.savePath

    # 폴더 생성
    def createFolder(self, directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
                self.PrintLog('Create folder')

        except OSError:
            self.PrintLog('Error : Createing directory', 'red')

    def DeleteHiddenSheet(self):
        sNames = self.wb.sheetnames
        if ('doc' in sNames):
            sNames.remove('doc')

        deleteDataList = []
        doOnce = True
        for idx, name in enumerate(sNames):
            tempWs = self.wb[name]
            if tempWs.sheet_state != 'visible':
                deleteDataList.append(name)
                if doOnce:
                    self.PrintLog('\nhidden Sheet list :', 'red')
                    doOnce = False
                self.PrintLog(name, 'red')

        for delName in deleteDataList:
            sNames.remove(delName)

        self.PrintLog("\nDo converting list :", 'blue')
        for idx in range(len(sNames)):
            self.PrintLog(sNames[idx], 'blue')
        self.PrintLog('', 'blue')

        return sNames

    def GetFileName(self):
        return self.fileName

    # 로딩바 값 설정 이벤트 바인드
    def SetLoadingbarEventBind(self, SetBar):
        self.SetLoadingbar_Handle = SetBar

    # 로그 생성 이벤트 바인드
    def PrintLogEventBind(self, print):
        self.PrintLog = print
    # 비어 있는 행, // or comment 로 구성된 열 제외

    def ExceptColumn(self):
        try:
            sheetNum = len(self.convertList)
            self.isTypename_Insheet_List.clear()

            self.newWbDatas = []
            for i, name in enumerate(self.convertList):
                ws = self.wb[name]

                cnt = 0
                for k in range(100):
                    if list(ws.rows)[k][0].value == None:
                        cnt += 1
                    else:
                        break

                if cnt > 0:
                    ws.delete_rows(1, cnt)

                row = list(ws.rows)[0]
                self.isTypename_Insheet_List.append(False)
                for type in DataType:
                    if DataType[type.name].value == row[0].value:
                        row = list(ws.rows)[1]
                        self.isTypename_Insheet_List[i] = True
                        break

                newWsDatas = []
                for idx, cell in enumerate(row):
                    if cell.value == None:
                        continue
                    if (not(self.del_ColName[0] in cell.value)) and (not(self.del_ColName[1] in cell.value)):
                        newWsDatas.append(list(ws.columns)[idx])

                    bar_Value = (i / sheetNum * 50) + (idx / len(row) * (sheetNum / 50))
                    self.SetLoadingbar_Handle(bar_Value)

                self.newWbDatas.append(newWsDatas)

            self.SetLoadingbar_Handle(50)
        except:
            self.PrintLog('error : 예외 처리중 에러 발생', 'red')
    # 데이터 타입, 데이터 None 찾아 로그 입력

    def PrintErrorLog(self):
        try:

            for sheetIdx, wsData in enumerate(self.newWbDatas):
                for colIdx, colData in enumerate(wsData):
                    try:
                        typeName = DataType(colData[0].value).value

                    except:
                        continue

                    for cellCount in range(len(colData)):
                        if cellCount < 2:
                            continue

                        dataType = SwitchType(typeName)

                        if colData[cellCount].value == None:
                            if dataType != str:
                                
                                if colIdx <= 25:
                                    rowcellName = chr(65 + colIdx)
                                else:
                                    remainder = colIdx % 25 - 1
                                    value = int(colIdx / 25) - 1
                                    rowcellName = '{}{}'.format(
                                        chr(65 + value), chr(65 + remainder))

                                self.PrintLog('>> warning:: sheet: {} cell: {}{} -> data empty'.format(
                                    self.convertList[sheetIdx], rowcellName, cellCount + 1), 'orange')
                                continue

                        currentCellType = GetType(colData[cellCount].value)
                        if currentCellType != dataType:
                            if dataType == int and currentCellType == str:
                                try:
                                    int(colData[cellCount].value)
                                except:
                                    if colIdx <= 25:
                                        rowcellName = chr(65 + colIdx)
                                    else:
                                        remainder = colIdx % 25 - 1
                                        value = int(colIdx / 25) - 1
                                        rowcellName = '{}{}'.format(
                                            chr(65 + value), chr(65 + remainder))

                                    self.PrintLog('>> error:: sheet: {} cell: {}{} -> different type'.format(
                                        self.convertList[sheetIdx], rowcellName, cellCount + 1), 'red')

                        if dataType == str:
                            if ',' in str(colData[cellCount].value):
                                if colIdx <= 25:
                                    rowcellName = chr(65 + colIdx)
                                else:
                                    remainder = colIdx % 25 - 1
                                    value = int(colIdx / 25) - 1
                                    rowcellName = '{}{}'.format(
                                        chr(65 + value), chr(65 + remainder))

                                self.PrintLog(">> error:: sheet: {} cell: {}{} -> Comprise ','".format(
                                    self.convertList[sheetIdx], rowcellName, cellCount + 1), 'red')
        except:
            self.PrintLog('error : 데이터 확인 중 에러 발생', 'red')

    # csv로 변환
    def ConversionToCsv(self):
        try:
            # csv로 변환해서 저장
            sheetNum = len(self.convertList)

            for idx, name in enumerate(self.convertList):
                try:
                    ob = csv.writer(open(self.savePath + '/' + name +
                                    '.csv', 'w', newline="", encoding='utf-8-sig'))
                except:
                    self.PrintLog('csv file access failure', 'red')
                    return False

                data = self.newWbDatas[idx]

                firstRowTransformation = False
                if (self.delFirstrow and self.isTypename_Insheet_List[idx]):
                    firstRowTransformation = True

                for j in range(len(data[0])):
                    row = []
                    if firstRowTransformation and j == 0:
                        continue

                    for i in data:
                        row.append(i[j].value)

                    ob.writerow(row)

                    bar_Value = (idx / sheetNum * 50) + (j / len(data[0]) * (sheetNum / 50)) + 50
                    self.SetLoadingbar_Handle(bar_Value)

           
        except:
            self.PrintLog('error : csv 변환 중 에러 발생', 'red')

    # json 으로 변환
    # def ConversionToJson(self):
    #     try:
    #         sheetNum = len(self.convertList)

    #         for idx, name in enumerate(self.convertList):
    #             data = self.newWbDatas[idx]
    #             data_list = []
    #             for row_num in range(len(data[0])):
    #                 if self.isTypename_Insheet_List[idx] and row_num <= 1:
    #                     continue
                    
    #                 tmp_dict = {}
    #                 for col_num in data:
    #                     if self.isTypename_Insheet_List[idx] == True:
    #                         tmp_dict[col_num[1].value] = col_num[row_num].value
    #                     else:
    #                         tmp_dict[col_num[0].value] = col_num[row_num].value

    #                 data_list.append(tmp_dict)

    #                 bar_Value = (idx / sheetNum * 50) + (row_num / len(data[0]) * (sheetNum / 50)) + 50
    #                 self.SetLoadingbar_Handle(bar_Value)

    #             try:
    #                 with open(self.savePath + '/' + name + '.json', 'w', encoding='utf-8') as fp:
    #                     json.dump(data_list, fp, indent=4, ensure_ascii=False)
                    
    #             except:
    #                 self.PrintLog('json file access failure', 'red')
    #                 return False

    #     except:
    #         self.PrintLog('error : json 변환 중 에러 발생', 'red')

    def Convert(self, kind):
        if not self.fileName:
            return False

        currentSaveTime = os.path.getmtime(self.fileName)
        if self.lastSaveTime != currentSaveTime:
            try:
                self.PrintLog('WorkBook reloading...')
                self.lastSaveTime = currentSaveTime
                self.wb.close()
                self.wb = openpyxl.load_workbook(str(self.fileName), data_only=True)
            except:
                self.PrintErrorLog('error : WorkBook relod failed', 'red')

        self.ExceptColumn()
        self.PrintErrorLog()
        # 지정된 경로 혹은 폴더 생성
        self.createFolder(self.savePath)
        if kind == 'csv':
            self.ConversionToCsv()
        # elif kind == 'json':
        #     self.ConversionToJson()

        for i in range(101):
            time.sleep(0.01)
            self.SetLoadingbar_Handle(90 + i / 10)
            if i >= 100:
                self.PrintLog('\nComplete file list :', 'blue')
                for j in (self.convertList):
                    self.PrintLog(j, 'blue')
                self.PrintLog('\nConversion Success!! try the open file button to click', 'blue')

        self.wb.close()
        self.completeConvert = True
        return True
